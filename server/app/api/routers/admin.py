import hashlib
import io
import json
import logging
import re
import secrets
import uuid
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File as FastAPIFile, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from app.auth import get_current_user
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import get_db
from app.models.area import Area
from app.models.catalog import CatalogLabelType, CatalogModel

logger = logging.getLogger(__name__)
from app.models.client import Client
from app.models.file import EntityFile, File
from app.models.plant import Plant
from app.models.policy import Policy, PolicyDelivery, PolicyDeliveryReport, PolicyPrinter, PolicyPrinterAssignment, PolicyVisit
from app.models.printer import Printer
from app.models.report import Report
from app.models.sync import SyncLog
from app.models.user import User

settings = get_settings()
from app.schemas.admin import (
    AreaCreate,
    AreaListItem,
    AssignPrintersRequest,
    AssignTechnicianRequest,
    CatalogModelCreate,
    CatalogModelItem,
    ClientCreate,
    ClientListItem,
    ClientUpdate,
    GenerateVisitsRequest,
    PlantCreate,
    PlantListItem,
    PlantUpdate,
    PolicyCreate,
    PolicyDeliveryCreate,
    PolicyDeliveryItem,
    PolicyDetail,
    PolicyListItem,
    PolicyPrinterAssignmentItem,
    PolicyPrinterItem,
    PolicyUpdate,
    PolicyVisitItem,
    PolicyVisitUpdate,
    PrinterCreate,
    PrinterListItem,
    PrinterUpdate,
    ReportDetail,
    ReportListItem,
    ReviewRequest,
    SyncHistoryItem,
    TechnicianCreate,
    TechnicianListItem,
    TechnicianUpdate,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])

# Checkbox keys from the Flutter app that indicate equipment damage
_DAMAGE_KEYS = {"Rodillo dañado", "Cabezal dañado", "Sensor ribbon dañado", "Sensor papel dañado"}


def _now_utc() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _policy_status(end_date: datetime) -> str:
    # Normalize to naive UTC: Pydantic v2 parses ISO strings with 'Z' as
    # timezone-aware datetimes; DB-loaded datetimes are naive. Stripping tzinfo
    # (replace, not astimezone) is safe because we always work in UTC.
    end = end_date.replace(tzinfo=None)
    now = _now_utc()
    if end < now:
        return "Expired"
    if end < now + timedelta(days=30):
        return "Expiring"
    return "Active"


def _printer_status(last_checkboxes_json: str | None) -> str:
    if not last_checkboxes_json:
        return "Sin Historial"
    try:
        checkboxes: dict = json.loads(last_checkboxes_json)
        for key in _DAMAGE_KEYS:
            if checkboxes.get(key) is True:
                return "En Atención"
        return "Correcto"
    except (json.JSONDecodeError, AttributeError):
        return "Sin Historial"


# ---------------------------------------------------------------------------
# GET /api/admin/reports
# ---------------------------------------------------------------------------

@router.get("/reports", response_model=dict)
def list_reports(
    client_id: str | None = Query(None),
    tech_id: str | None = Query(None),
    status: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    """List reports with optional filters and pagination."""
    q = (
        db.query(
            Report.id,
            Report.code,
            Report.service_type,
            Report.status,
            Report.service_date,
            Report.sync_date,
            Printer.serial_number.label("printer_serial"),
            Client.name.label("client_name"),
            User.name.label("tech_name"),
        )
        .outerjoin(Printer, Report.printer_id == Printer.id)
        .outerjoin(Client, Printer.client_id == Client.id)
        .outerjoin(User, Report.tech_id == User.id)
    )

    if client_id:
        q = q.filter(Client.id == client_id)
    if tech_id:
        q = q.filter(Report.tech_id == tech_id)
    if status:
        q = q.filter(Report.status == status)
    if date_from:
        q = q.filter(Report.service_date >= date_from)
    if date_to:
        q = q.filter(Report.service_date <= date_to)

    total = q.count()
    rows = q.order_by(Report.service_date.desc()).offset(offset).limit(limit).all()

    items = [
        ReportListItem(
            id=r.id,
            code=r.code,
            service_type=r.service_type,
            status=r.status,
            service_date=r.service_date,
            sync_date=r.sync_date,
            printer_serial=r.printer_serial,
            client_name=r.client_name,
            tech_name=r.tech_name,
        )
        for r in rows
    ]

    return {"total": total, "offset": offset, "limit": limit, "items": [i.model_dump() for i in items]}


# ---------------------------------------------------------------------------
# GET /api/admin/reports/{report_id}
# ---------------------------------------------------------------------------

@router.get("/reports/{report_id}", response_model=ReportDetail)
def get_report(report_id: str, db: Session = Depends(get_db)) -> ReportDetail:
    """Get full detail of a single report."""
    row = (
        db.query(
            Report,
            Printer.serial_number.label("printer_serial"),
            Printer.code.label("printer_code"),
            Client.name.label("client_name"),
            User.name.label("tech_name"),
            User.code.label("tech_code"),
        )
        .outerjoin(Printer, Report.printer_id == Printer.id)
        .outerjoin(Client, Printer.client_id == Client.id)
        .outerjoin(User, Report.tech_id == User.id)
        .filter(Report.id == report_id)
        .first()
    )

    if not row:
        raise HTTPException(status_code=404, detail="Report not found")

    report, printer_serial, printer_code, client_name, tech_name, tech_code = row

    return ReportDetail(
        id=report.id,
        code=report.code,
        printer_id=report.printer_id,
        tech_id=report.tech_id,
        service_type=report.service_type,
        status=report.status,
        service_date=report.service_date,
        linear_inches_counter=report.linear_inches_counter,
        darkness_level=report.darkness_level,
        label_type_id=report.label_type_id,
        technical_checkboxes=report.technical_checkboxes,
        notes=report.notes,
        signature_name=report.signature_name,
        signature_role=report.signature_role,
        signature_image_path=report.signature_image_path,
        photo_paths=report.photo_paths,
        photo_count=report.photo_count,
        internal_notes=report.internal_notes,
        sync_date=report.sync_date,
        created_at=report.created_at,
        printer_serial=printer_serial,
        printer_code=printer_code,
        client_name=client_name,
        tech_name=tech_name,
        tech_code=tech_code,
    )


# ---------------------------------------------------------------------------
# POST /api/admin/reports/{report_id}/review
# ---------------------------------------------------------------------------

@router.post("/reports/{report_id}/review")
def review_report(
    report_id: str,
    body: ReviewRequest,
    db: Session = Depends(get_db),
) -> dict:
    """Approve or reject a report and optionally add internal notes."""
    report: Report | None = db.get(Report, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if body.status not in ("approved", "rejected"):
        raise HTTPException(status_code=422, detail="status must be 'approved' or 'rejected'")

    report.status = "Reviewed-" + body.status.capitalize()
    if body.notes:
        report.internal_notes = body.notes

    db.commit()
    return {"success": True, "id": report_id, "status": report.status}


# ---------------------------------------------------------------------------
# GET /api/admin/reports/{report_id}/files
# ---------------------------------------------------------------------------

@router.get("/reports/{report_id}/files")
def get_report_files(report_id: str, db: Session = Depends(get_db)) -> dict:
    """Return URLs for photos, signature and PDF associated with a report."""
    rows = (
        db.query(EntityFile, File)
        .join(File, EntityFile.file_id == File.id)
        .filter(
            EntityFile.entity_id == report_id,
            EntityFile.entity_type == "report",
        )
        .all()
    )

    upload_dir_path = Path(settings.upload_dir)
    photos: list[str] = []
    signature: str | None = None
    pdf: str | None = None

    for ef, f in rows:
        try:
            rel = Path(f.storage_path).relative_to(upload_dir_path)
            url = f"/uploads/{rel.as_posix()}"
        except ValueError:
            url = f"/uploads/{f.storage_path.lstrip('/')}"
        if ef.file_category == "photo":
            photos.append(url)
        elif ef.file_category == "signature":
            signature = url
        elif ef.file_category == "pdf":
            pdf = url

    return {"photos": photos, "signature": signature, "pdf": pdf}


# ---------------------------------------------------------------------------
# GET /api/admin/clients
# ---------------------------------------------------------------------------

@router.get("/clients", response_model=dict)
def list_clients(
    search: str | None = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    """List clients with plant/printer/policy counts."""
    plant_count_sq = (
        db.query(Plant.client_id, func.count(Plant.id).label("cnt"))
        .group_by(Plant.client_id)
        .subquery()
    )
    printer_count_sq = (
        db.query(Printer.client_id, func.count(Printer.id).label("cnt"))
        .group_by(Printer.client_id)
        .subquery()
    )
    policy_count_sq = (
        db.query(Policy.client_id, func.count(Policy.id).label("cnt"))
        .filter(Policy.status == "Active")
        .group_by(Policy.client_id)
        .subquery()
    )

    q = (
        db.query(
            Client,
            func.coalesce(plant_count_sq.c.cnt, 0).label("plant_count"),
            func.coalesce(printer_count_sq.c.cnt, 0).label("printer_count"),
            func.coalesce(policy_count_sq.c.cnt, 0).label("active_policy_count"),
        )
        .outerjoin(plant_count_sq, Client.id == plant_count_sq.c.client_id)
        .outerjoin(printer_count_sq, Client.id == printer_count_sq.c.client_id)
        .outerjoin(policy_count_sq, Client.id == policy_count_sq.c.client_id)
    )

    if search:
        q = q.filter(Client.name.ilike(f"%{search}%") | Client.rfc.ilike(f"%{search}%"))

    total = q.count()
    rows = q.order_by(Client.name).offset(offset).limit(limit).all()

    items = [
        ClientListItem(
            id=client.id,
            name=client.name,
            rfc=client.rfc,
            address=client.address,
            is_active=client.is_active,
            plant_count=plant_count,
            printer_count=printer_count,
            active_policy_count=active_policy_count,
        )
        for client, plant_count, printer_count, active_policy_count in rows
    ]

    return {"total": total, "offset": offset, "limit": limit, "items": [i.model_dump() for i in items]}


# ---------------------------------------------------------------------------
# GET /api/admin/clients/{client_id}/detail
# ---------------------------------------------------------------------------

@router.get("/clients/{client_id}/detail", response_model=dict)
def get_client_detail(client_id: str, db: Session = Depends(get_db)) -> dict:
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    plants = db.query(Plant).filter(Plant.client_id == client_id).all()
    printers = db.query(Printer).filter(Printer.client_id == client_id).all()

    now = _now_utc()
    thirty_days_ago = now - timedelta(days=30)

    en_atencion_count = 0
    printer_rows = []

    for p in printers:
        last_report = (
            db.query(Report)
            .filter(Report.printer_id == p.id)
            .order_by(Report.service_date.desc())
            .first()
        )
        _en_atencion = False
        ultimo_contador = None
        if last_report:
            ultimo_contador = last_report.linear_inches_counter
            try:
                cb = json.loads(last_report.technical_checkboxes or "{}")
                for key in _DAMAGE_KEYS:
                    if cb.get(key) is True:
                        _en_atencion = True
                        break
            except (json.JSONDecodeError, AttributeError):
                pass

        if _en_atencion:
            en_atencion_count += 1

        total_reportes = (
            db.query(func.count(Report.id)).filter(Report.printer_id == p.id).scalar() or 0
        )
        model = db.get(CatalogModel, p.model_id) if p.model_id else None
        plant_obj = db.get(Plant, p.plant_id) if p.plant_id else None
        area_obj = db.get(Area, p.area_id) if p.area_id else None
        model_name = f"{model.brand} {model.model_name}" if model else None

        printer_rows.append({
            "id": p.id,
            "code": p.code,
            "serial_number": p.serial_number,
            "is_active": p.is_active,
            "model_name": model_name,
            "area_name": area_obj.name if area_obj else None,
            "plant_name": plant_obj.name if plant_obj else None,
            "ultimo_contador": ultimo_contador,
            "en_atencion": _en_atencion,
            "total_reportes": total_reportes,
        })

    # Reports last 30 days
    reports_30d = (
        db.query(Report)
        .join(Printer, Report.printer_id == Printer.id)
        .filter(Printer.client_id == client_id)
        .filter(Report.service_date >= thirty_days_ago)
        .all()
    )
    r_total = len(reports_30d)
    r_preventivos = sum(1 for r in reports_30d if (r.service_type or "").lower() == "preventivo")
    r_correctivos = sum(1 for r in reports_30d if (r.service_type or "").lower() == "correctivo")
    r_diagnosticos = sum(
        1 for r in reports_30d
        if (r.service_type or "").lower() in ("diagnóstico", "diagnostico", "diagnóstico")
    )

    # Policies
    all_policies = db.query(Policy).filter(Policy.client_id == client_id).all()
    polizas_activas = sum(1 for p in all_policies if p.end_date and p.end_date.replace(tzinfo=None) >= now)
    polizas_vencidas = sum(1 for p in all_policies if p.end_date and p.end_date.replace(tzinfo=None) < now)

    # Top printer (most reports)
    impresora_mas_servicios = None
    if printer_rows:
        top = max(printer_rows, key=lambda x: x["total_reportes"])
        if top["total_reportes"] > 0:
            impresora_mas_servicios = {
                "serial_number": top["serial_number"],
                "code": top["code"],
                "model_name": top["model_name"],
                "total_reportes": top["total_reportes"],
            }

    logo_url = f"/uploads/{client.logo_path}" if client.logo_path else None

    return {
        "client": {
            "id": client.id,
            "name": client.name,
            "rfc": client.rfc,
            "address": client.address,
            "is_active": client.is_active,
            "logo_url": logo_url,
        },
        "plants": [
            {"id": pl.id, "name": pl.name, "contact_name": pl.contact_name, "phone": pl.phone}
            for pl in plants
        ],
        "stats": {
            "total_impresoras": len(printers),
            "impresoras_activas": sum(1 for p in printers if p.is_active),
            "impresoras_en_atencion": en_atencion_count,
            "reportes_ultimo_mes": {
                "total": r_total,
                "preventivos": r_preventivos,
                "correctivos": r_correctivos,
                "diagnosticos": r_diagnosticos,
            },
            "polizas_activas": polizas_activas,
            "polizas_vencidas": polizas_vencidas,
            "impresora_mas_servicios": impresora_mas_servicios,
        },
        "printers": printer_rows,
    }


# ---------------------------------------------------------------------------
# PATCH /api/admin/clients/{client_id}/logo
# DELETE /api/admin/clients/{client_id}/logo
# ---------------------------------------------------------------------------

_LOGO_ALLOWED_TYPES = {"image/png", "image/jpeg"}
_LOGO_MAX_BYTES = 2 * 1024 * 1024  # 2 MB
_LOGO_MAGIC = {
    "image/png": b"\x89PNG",
    "image/jpeg": b"\xff\xd8",
}


@router.patch("/clients/{client_id}/logo", response_model=dict)
async def upload_client_logo(
    client_id: str,
    logo: UploadFile = FastAPIFile(...),
    db: Session = Depends(get_db),
    _current_user: dict = Depends(get_current_user),
) -> dict:
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    content_type = (logo.content_type or "").split(";")[0].strip()
    if content_type not in _LOGO_ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes PNG o JPG")

    content = await logo.read()
    if len(content) > _LOGO_MAX_BYTES:
        raise HTTPException(status_code=400, detail="El archivo no puede superar 2 MB")

    magic = _LOGO_MAGIC[content_type]
    if not content[: len(magic)] == magic:
        raise HTTPException(status_code=400, detail="El contenido del archivo no coincide con el tipo declarado")

    ext = "png" if content_type == "image/png" else "jpg"
    logos_dir = Path(settings.upload_dir) / "logos"
    logos_dir.mkdir(parents=True, exist_ok=True)

    # Remove previous logo (either extension) if it exists
    for old_ext in ("png", "jpg"):
        old_path = logos_dir / f"{client_id}.{old_ext}"
        if old_path.exists():
            old_path.unlink()

    file_path = logos_dir / f"{client_id}.{ext}"
    file_path.write_bytes(content)

    client.logo_path = f"logos/{client_id}.{ext}"
    db.commit()

    return {"logo_url": f"/uploads/logos/{client_id}.{ext}"}


@router.delete("/clients/{client_id}/logo", response_model=dict)
def delete_client_logo(
    client_id: str,
    db: Session = Depends(get_db),
    _current_user: dict = Depends(get_current_user),
) -> dict:
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    if client.logo_path:
        file_path = Path(settings.upload_dir) / client.logo_path
        if file_path.exists():
            file_path.unlink()
        client.logo_path = None
        db.commit()

    return {"ok": True}


# ---------------------------------------------------------------------------
# GET /api/admin/technicians
# ---------------------------------------------------------------------------

@router.get("/technicians", response_model=dict)
def list_technicians(
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    """List technicians with their report count and last sync."""
    report_count_sq = (
        db.query(Report.tech_id, func.count(Report.id).label("cnt"))
        .group_by(Report.tech_id)
        .subquery()
    )

    q = (
        db.query(
            User,
            func.coalesce(report_count_sq.c.cnt, 0).label("reports_count"),
        )
        .outerjoin(report_count_sq, User.id == report_count_sq.c.tech_id)
        .filter(User.role == "technician")
    )

    total = q.count()
    rows = q.order_by(User.name).offset(offset).limit(limit).all()

    items = [
        TechnicianListItem(
            id=user.id,
            code=user.code,
            name=user.name,
            email=user.email,
            role=user.role,
            reports_count=reports_count,
            last_sync_at=user.last_sync_at,
        )
        for user, reports_count in rows
    ]

    return {"total": total, "offset": offset, "limit": limit, "items": [i.model_dump() for i in items]}


# ---------------------------------------------------------------------------
# GET /api/admin/technicians/{tech_id}  — Perfil
# ---------------------------------------------------------------------------

@router.get("/technicians/{tech_id}", response_model=dict)
def get_technician_detail(tech_id: str, db: Session = Depends(get_db)) -> dict:
    tech = db.query(User).filter(User.id == tech_id, User.role == "technician").first()
    if not tech:
        raise HTTPException(status_code=404, detail="Technician not found")
    report_count = db.query(Report).filter(Report.tech_id == tech_id).count()
    signature_url = f"/uploads/{tech.signature_path}" if tech.signature_path else None
    return {
        "id": tech.id, "code": tech.code, "name": tech.name,
        "email": tech.email, "role": tech.role, "is_active": tech.is_active,
        "signature_url": signature_url,
        "last_sync_at": tech.last_sync_at.isoformat() if tech.last_sync_at else None,
        "report_count": report_count,
    }


# ---------------------------------------------------------------------------
# GET /api/admin/technicians/{tech_id}/reports  — Reportes con filtro de fecha
# ---------------------------------------------------------------------------

@router.get("/technicians/{tech_id}/reports", response_model=dict)
def get_technician_reports(
    tech_id: str,
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    q = (
        db.query(Report)
        .filter(Report.tech_id == tech_id)
        .order_by(Report.service_date.desc())
    )
    if date_from:
        q = q.filter(Report.service_date >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.filter(Report.service_date <= datetime.fromisoformat(date_to))
    total = q.count()
    rows = q.offset(offset).limit(limit).all()
    items = []
    for r in rows:
        printer = db.get(Printer, r.printer_id) if r.printer_id else None
        client = db.get(Client, printer.client_id) if printer and printer.client_id else None
        items.append({
            "id": r.id, "code": r.code, "service_type": r.service_type,
            "service_date": r.service_date.isoformat() if r.service_date else None,
            "status": r.status,
            "printer_serial": printer.serial_number if printer else None,
            "client_name": client.name if client else None,
        })
    return {"total": total, "items": items}


# ---------------------------------------------------------------------------
# GET /api/admin/printers
# ---------------------------------------------------------------------------

@router.get("/printers", response_model=dict)
def list_printers(
    client_id: str | None = Query(None),
    printer_status: str | None = Query(None),
    search: str | None = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    """List printers with their last service status, model info, and optional search."""
    # Subquery: latest report per printer
    latest_report_sq = (
        db.query(
            Report.printer_id,
            func.max(Report.service_date).label("max_date"),
        )
        .group_by(Report.printer_id)
        .subquery()
    )
    latest_full_sq = (
        db.query(Report.printer_id, Report.service_date, Report.technical_checkboxes)
        .join(
            latest_report_sq,
            (Report.printer_id == latest_report_sq.c.printer_id)
            & (Report.service_date == latest_report_sq.c.max_date),
        )
        .subquery()
    )

    q = (
        db.query(
            Printer,
            Client.name.label("client_name"),
            Plant.name.label("plant_name"),
            Area.name.label("area_name"),
            CatalogModel.brand.label("model_brand"),
            CatalogModel.model_name.label("model_name"),
            CatalogModel.dpi.label("model_dpi"),
            latest_full_sq.c.service_date.label("last_service_date"),
            latest_full_sq.c.technical_checkboxes.label("last_checkboxes"),
        )
        .outerjoin(Client, Printer.client_id == Client.id)
        .outerjoin(Plant, Printer.plant_id == Plant.id)
        .outerjoin(Area, Printer.area_id == Area.id)
        .outerjoin(CatalogModel, Printer.model_id == CatalogModel.id)
        .outerjoin(latest_full_sq, Printer.id == latest_full_sq.c.printer_id)
    )

    if client_id:
        q = q.filter(Printer.client_id == client_id)
    if search:
        pattern = f"%{search}%"
        q = q.filter(
            Printer.serial_number.ilike(pattern)
            | Printer.code.ilike(pattern)
            | CatalogModel.model_name.ilike(pattern)
            | CatalogModel.brand.ilike(pattern)
        )

    rows = q.order_by(Printer.serial_number).offset(offset).limit(limit).all()
    total = q.count()

    items = []
    for printer, client_name, plant_name, area_name, model_brand, model_name, model_dpi, last_service_date, last_checkboxes in rows:
        status = _printer_status(last_checkboxes)
        if printer_status and status != printer_status:
            continue
        items.append(
            PrinterListItem(
                id=printer.id,
                code=printer.code,
                serial_number=printer.serial_number,
                client_name=client_name,
                plant_name=plant_name,
                area_name=area_name,
                model_brand=model_brand,
                model_name=model_name,
                model_dpi=model_dpi,
                last_service_date=last_service_date,
                printer_status=status,
            )
        )

    return {"total": total, "offset": offset, "limit": limit, "items": [i.model_dump() for i in items]}


# ---------------------------------------------------------------------------
# GET /api/admin/printers/template/download  — Excel template
# MUST be registered before /printers/{printer_id} to avoid route capture
# ---------------------------------------------------------------------------

_SAFE_STR_RE = re.compile(r"[<>;/*\x00]|--|/\*|\*/")


def _sanitize_upload_str(s: str) -> str:
    return _SAFE_STR_RE.sub("", s).strip()


@router.get("/printers/template/download")
def download_printers_template(db: Session = Depends(get_db)) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # Sheet 1 — Impresoras (data entry)
    ws = wb.active
    ws.title = "Impresoras"
    header = ["serie", "cliente", "planta", "area", "marca", "modelo", "dpi"]
    ws.append(header)
    for col, _ in enumerate(header, start=1):
        ws.cell(1, col).font = Font(bold=True)
    ws.append(["ZBR-12345", "Mi Cliente S.A.", "Planta Norte", "Almacén", "Zebra", "ZT411", "203"])

    # Sheet 2 — Instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 12
    rows_inst = [
        ("Campo", "Descripción", "Ejemplo", "Obligatorio"),
        ("serie", "Número de serie de la impresora (debe ser único)", "ZBR-12345", "Sí"),
        ("cliente", "Nombre exacto del cliente registrado en el sistema", "Mi Cliente S.A.", "Sí"),
        ("planta", "Nombre de la planta (se crea automáticamente si no existe)", "Planta Norte", "Sí"),
        ("area", "Nombre del área (se crea automáticamente si no existe)", "Almacén", "Sí"),
        ("marca", "Marca de la impresora (se crea modelo si no existe)", "Zebra", "No"),
        ("modelo", "Nombre del modelo de la impresora", "ZT411", "No"),
        ("dpi", "Resolución en DPI (default: 203 si se omite)", "203", "No"),
    ]
    for r in rows_inst:
        ws2.append(r)
    for col in range(1, 5):
        ws2.cell(1, col).font = Font(bold=True)

    # Sheet 3 — Catálogos
    ws3 = wb.create_sheet("Catálogos")
    ws3.append(["CLIENTES ACTIVOS", "", "", "MODELOS ACTIVOS", "", ""])
    ws3.append(["Nombre", "", "", "Marca", "Modelo", "DPI"])
    for col in [1, 4]:
        ws3.cell(1, col).font = Font(bold=True)
        ws3.cell(2, col).font = Font(bold=True)

    clients_q = db.query(Client).filter(Client.is_active.is_(True)).order_by(Client.name).all()
    models_q = db.query(CatalogModel).order_by(CatalogModel.brand, CatalogModel.model_name).all()
    max_rows = max(len(clients_q), len(models_q), 1)
    for i in range(max_rows):
        c_name = clients_q[i].name if i < len(clients_q) else ""
        m_brand = models_q[i].brand if i < len(models_q) else ""
        m_model = models_q[i].model_name if i < len(models_q) else ""
        m_dpi = str(models_q[i].dpi) if i < len(models_q) else ""
        ws3.append([c_name, "", "", m_brand, m_model, m_dpi])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_impresoras.xlsx"},
    )


# ---------------------------------------------------------------------------
# POST /api/admin/printers/bulk-upload
# ---------------------------------------------------------------------------

@router.post("/printers/bulk-upload", response_model=dict)
async def bulk_upload_printers(
    file: UploadFile = FastAPIFile(...),
    db: Session = Depends(get_db),
) -> dict:
    import csv
    import openpyxl

    filename = file.filename or ""
    ext = Path(filename).suffix.lower()
    if ext not in (".xlsx", ".csv"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .csv")

    content = await file.read()

    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo no puede superar 5 MB")

    if ext == ".xlsx" and content[:4] != b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx)")

    rows: list[dict[str, str]] = []
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] | None = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip().lower() if c is not None else "" for c in row]
                continue
            if all(c is None for c in row):
                continue
            rows.append(dict(zip(headers, [str(c).strip() if c is not None else "" for c in row])))
        wb.close()
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): v.strip() for k, v in row.items()})

    if len(rows) > 500:
        raise HTTPException(status_code=400, detail="El archivo no puede tener más de 500 filas")

    total = len(rows)
    exitosas = 0
    errores: list[dict] = []

    for i, row in enumerate(rows, start=2):
        serie = _sanitize_upload_str(row.get("serie", ""))
        cliente_name = _sanitize_upload_str(row.get("cliente", ""))
        planta_name = _sanitize_upload_str(row.get("planta", ""))
        area_name = _sanitize_upload_str(row.get("area", ""))
        marca = _sanitize_upload_str(row.get("marca", ""))
        modelo = _sanitize_upload_str(row.get("modelo", ""))
        dpi_str = _sanitize_upload_str(row.get("dpi", "203"))

        try:
            if not serie:
                raise ValueError("Serie requerida")

            existing = db.query(Printer).filter(Printer.serial_number == serie).first()
            if existing:
                raise ValueError(f"Serie '{serie}' ya existe")

            if not cliente_name:
                raise ValueError("Cliente requerido")

            client_obj = (
                db.query(Client)
                .filter(func.lower(Client.name) == func.lower(cliente_name))
                .first()
            )
            if not client_obj:
                raise ValueError(f"Cliente '{cliente_name}' no encontrado")

            plant_obj = None
            if planta_name:
                plant_obj = (
                    db.query(Plant)
                    .filter(
                        Plant.client_id == client_obj.id,
                        func.lower(Plant.name) == func.lower(planta_name),
                    )
                    .first()
                )
                if not plant_obj:
                    plant_obj = Plant(id=str(uuid.uuid4()), client_id=client_obj.id, name=planta_name)
                    db.add(plant_obj)
                    db.flush()

            area_obj = None
            if area_name and plant_obj:
                area_obj = (
                    db.query(Area)
                    .filter(
                        Area.plant_id == plant_obj.id,
                        func.lower(Area.name) == func.lower(area_name),
                    )
                    .first()
                )
                if not area_obj:
                    area_obj = Area(id=str(uuid.uuid4()), plant_id=plant_obj.id, name=area_name)
                    db.add(area_obj)
                    db.flush()

            model_obj = None
            if marca or modelo:
                dpi_val = int(dpi_str) if dpi_str.isdigit() else 203
                model_obj = (
                    db.query(CatalogModel)
                    .filter(
                        func.lower(CatalogModel.brand) == func.lower(marca),
                        func.lower(CatalogModel.model_name) == func.lower(modelo),
                    )
                    .first()
                )
                if not model_obj:
                    model_obj = CatalogModel(
                        id=str(uuid.uuid4()),
                        brand=marca,
                        model_name=modelo,
                        dpi=dpi_val,
                    )
                    db.add(model_obj)
                    db.flush()

            code = _next_code(db, Printer, "I")
            new_printer = Printer(
                id=str(uuid.uuid4()),
                code=code,
                serial_number=serie,
                client_id=client_obj.id,
                plant_id=plant_obj.id if plant_obj else None,
                area_id=area_obj.id if area_obj else None,
                model_id=model_obj.id if model_obj else None,
                qr_uuid=str(uuid.uuid4()),
                is_active=True,
            )
            db.add(new_printer)
            db.commit()
            exitosas += 1

        except ValueError as exc:
            db.rollback()
            errores.append({"fila": i, "serie": serie or "—", "error": str(exc)})
            logger.warning("Bulk upload row %d error: %s", i, exc)
        except Exception as exc:
            db.rollback()
            errores.append({"fila": i, "serie": serie or "—", "error": "Error interno al procesar la fila"})
            logger.error("Bulk upload row %d unexpected error: %s", i, exc)

    return {"total": total, "exitosas": exitosas, "errores": errores}


# ---------------------------------------------------------------------------
# GET /api/admin/printers/{printer_id}  — Detalle
# ---------------------------------------------------------------------------

@router.get("/printers/{printer_id}", response_model=dict)
def get_printer_detail(printer_id: str, db: Session = Depends(get_db)) -> dict:
    printer = db.get(Printer, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")
    client = db.get(Client, printer.client_id) if printer.client_id else None
    plant = db.get(Plant, printer.plant_id) if printer.plant_id else None
    area = db.get(Area, printer.area_id) if printer.area_id else None
    model = db.get(CatalogModel, printer.model_id) if printer.model_id else None
    return {
        "id": printer.id,
        "code": printer.code,
        "serial_number": printer.serial_number,
        "qr_uuid": printer.qr_uuid,
        "is_active": printer.is_active,
        "client": {"id": client.id, "name": client.name} if client else None,
        "plant": {
            "id": plant.id, "name": plant.name,
            "contact_name": plant.contact_name, "phone": plant.phone,
        } if plant else None,
        "area": {"id": area.id, "name": area.name} if area else None,
        "model": {
            "id": model.id, "brand": model.brand,
            "model_name": model.model_name, "dpi": model.dpi,
        } if model else None,
    }


# ---------------------------------------------------------------------------
# GET /api/admin/printers/{printer_id}/reports  — Historial paginado
# ---------------------------------------------------------------------------

@router.get("/printers/{printer_id}/reports", response_model=dict)
def get_printer_reports(
    printer_id: str,
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    q = (
        db.query(Report)
        .filter(Report.printer_id == printer_id)
        .order_by(Report.service_date.desc())
    )
    total = q.count()
    rows = q.offset(offset).limit(limit).all()
    items = []
    for r in rows:
        tech = db.get(User, r.tech_id) if r.tech_id else None
        items.append({
            "id": r.id, "code": r.code, "service_type": r.service_type,
            "service_date": r.service_date.isoformat() if r.service_date else None,
            "status": r.status, "tech_name": tech.name if tech else None,
            "notes": r.notes, "signature_name": r.signature_name,
            "signature_role": r.signature_role,
            "linear_inches_counter": r.linear_inches_counter,
            "darkness_level": r.darkness_level,
            "technical_checkboxes": r.technical_checkboxes,
            "photo_count": r.photo_count or 0,
        })
    return {"total": total, "items": items}


# ---------------------------------------------------------------------------
# GET /api/admin/printers/{printer_id}/stats
# ---------------------------------------------------------------------------

@router.get("/printers/{printer_id}/stats", response_model=dict)
def get_printer_stats(printer_id: str, db: Session = Depends(get_db)) -> dict:
    thirty_days_ago = _now_utc() - timedelta(days=30)

    recent_reports = (
        db.query(Report)
        .filter(Report.printer_id == printer_id)
        .filter(Report.service_date >= thirty_days_ago)
        .order_by(Report.service_date.desc())
        .all()
    )
    last_report = (
        db.query(Report)
        .filter(Report.printer_id == printer_id)
        .order_by(Report.service_date.desc())
        .first()
    )

    contador_promedio: int | None = None
    if recent_reports:
        counters = [r.linear_inches_counter for r in recent_reports if r.linear_inches_counter is not None]
        if counters:
            contador_promedio = int(sum(counters) / len(counters))

    ultimo_contador: int | None = last_report.linear_inches_counter if last_report else None

    oscuridad_promedio: int | None = None
    if recent_reports:
        levels = [r.darkness_level for r in recent_reports if r.darkness_level is not None]
        if levels:
            oscuridad_promedio = int(sum(levels) / len(levels))

    etiqueta_frecuente: str | None = None
    if recent_reports:
        label_counts: dict[str, int] = {}
        for r in recent_reports:
            if r.label_type_id:
                label_counts[r.label_type_id] = label_counts.get(r.label_type_id, 0) + 1
        if label_counts:
            top_label_id = max(label_counts, key=lambda k: label_counts[k])
            label = db.get(CatalogLabelType, top_label_id)
            if label:
                etiqueta_frecuente = label.name

    ultima_observacion: str | None = last_report.notes if last_report else None

    _WARNING_KEYS = ["Rodillo dañado", "Cabezal dañado", "Sensor papel dañado", "Otros"]
    advertencias_activas: list[str] = []
    if last_report and last_report.technical_checkboxes:
        try:
            cb = json.loads(last_report.technical_checkboxes)
            for key in _WARNING_KEYS:
                if cb.get(key) is True:
                    advertencias_activas.append(key)
        except (json.JSONDecodeError, AttributeError):
            pass

    return {
        "contador_promedio": contador_promedio,
        "ultimo_contador": ultimo_contador,
        "oscuridad_promedio": oscuridad_promedio,
        "etiqueta_frecuente": etiqueta_frecuente,
        "ultima_observacion": ultima_observacion,
        "advertencias_activas": advertencias_activas,
    }


# ---------------------------------------------------------------------------
# GET /api/admin/policies
# ---------------------------------------------------------------------------

@router.get("/policies", response_model=dict)
def list_policies(
    client_id: str | None = Query(None),
    status: str | None = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> dict:
    """List policies with computed status and printer count."""
    printer_count_sq = (
        db.query(PolicyPrinter.policy_id, func.count(PolicyPrinter.id).label("cnt"))
        .group_by(PolicyPrinter.policy_id)
        .subquery()
    )

    q = (
        db.query(
            Policy,
            Client.name.label("client_name"),
            func.coalesce(printer_count_sq.c.cnt, 0).label("printer_count"),
        )
        .join(Client, Policy.client_id == Client.id)
        .outerjoin(printer_count_sq, Policy.id == printer_count_sq.c.policy_id)
    )

    if client_id:
        q = q.filter(Policy.client_id == client_id)

    rows = q.order_by(Policy.end_date).offset(offset).limit(limit).all()
    total = q.count()

    items = []
    for policy, client_name, printer_count in rows:
        computed_status = _policy_status(policy.end_date)
        if status and computed_status != status:
            continue
        items.append(
            PolicyListItem(
                id=policy.id,
                code=policy.code,
                folio=policy.folio,
                client_name=client_name,
                coverage_type=policy.coverage_type,
                start_date=policy.start_date,
                end_date=policy.end_date,
                status=computed_status,
                printer_count=printer_count,
                sla_notes=policy.sla_notes,
                frequency_maintenance=policy.frequency_maintenance,
            )
        )

    return {"total": total, "offset": offset, "limit": limit, "items": [i.model_dump() for i in items]}


# ---------------------------------------------------------------------------
# GET /api/admin/sync/history
# ---------------------------------------------------------------------------

def _enrich_sync_item(row: SyncLog, db: Session) -> tuple[str | None, str | None]:
    """Return (tech_name, detalle) for a SyncLog row by looking up the linked report."""
    if row.entity_type != "report":
        return None, None
    report = db.get(Report, row.entity_id)
    if not report:
        return None, None
    detalle = report.code
    tech = db.get(User, report.tech_id) if report.tech_id else None
    return (tech.name if tech else None), detalle


@router.get("/sync/history", response_model=dict)
def sync_history(
    status: str | None = Query(None),
    entity_type: str | None = Query(None),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
) -> dict:
    """Paginated sync log history."""
    q = db.query(SyncLog)

    if status:
        q = q.filter(SyncLog.status == status)
    if entity_type:
        q = q.filter(SyncLog.entity_type == entity_type)

    total = q.count()
    rows = q.order_by(SyncLog.synced_at.desc()).offset(offset).limit(limit).all()

    items = []
    for row in rows:
        tech_name, detalle = _enrich_sync_item(row, db)
        items.append(
            SyncHistoryItem(
                id=row.id,
                entity_type=row.entity_type,
                entity_id=row.entity_id,
                action=row.action,
                status=row.status,
                error_message=row.error_message,
                synced_at=row.synced_at,
                server_response=row.server_response,
                tech_name=tech_name,
                detalle=detalle,
            )
        )

    return {"total": total, "offset": offset, "limit": limit, "items": [i.model_dump() for i in items]}


# ===========================================================================
# Shared helpers
# ===========================================================================

def _next_code(db: Session, model, prefix: str, digits: int = 4) -> str:
    """Generate the next incremental readable code like C-0001, I-0001..."""
    pattern = f"{prefix}-%"
    rows = db.query(model.code).filter(model.code.like(pattern)).all()
    nums = []
    for (code,) in rows:
        if code:
            try:
                nums.append(int(code[len(prefix) + 1:]))
            except ValueError:
                pass
    return f"{prefix}-{max(nums, default=0) + 1:0{digits}d}"


_SPECIAL_CHARS = set("!@#$%^&*()_+-=[]{}|;':\",./<>?`~\\")


def _validate_password(password: str, username: str) -> None:
    """Raise HTTP 400 if password doesn't meet security rules."""
    if len(password) < 10:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 10 caracteres.")
    if not any(c.isupper() for c in password):
        raise HTTPException(status_code=400, detail="La contraseña debe contener al menos una letra mayúscula.")
    if not any(c.islower() for c in password):
        raise HTTPException(status_code=400, detail="La contraseña debe contener al menos una letra minúscula.")
    if not any(c.isdigit() for c in password):
        raise HTTPException(status_code=400, detail="La contraseña debe contener al menos un número.")
    if not any(c in _SPECIAL_CHARS for c in password):
        raise HTTPException(status_code=400, detail="La contraseña debe contener al menos un carácter especial (!@#$%^&*...).")
    if username and username.lower() in password.lower():
        raise HTTPException(status_code=400, detail="La contraseña no puede contener el nombre del usuario.")


def _hash_password(password: str) -> str:
    """Hash a password using PBKDF2-HMAC-SHA256 with a random salt."""
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 260_000)
    return f"{salt}${dk.hex()}"


# ===========================================================================
# FASE 3 — Gestión de Pólizas
# ===========================================================================

def _next_policy_code(db: Session) -> str:
    """Generate the next incremental code P-001, P-002..."""
    rows = db.query(Policy.code).filter(Policy.code.like("P-%")).all()
    nums = []
    for (code,) in rows:
        if code:
            try:
                nums.append(int(code[2:]))
            except ValueError:
                pass
    return f"P-{max(nums, default=0) + 1:03d}"


def _policy_detail(policy: Policy, db: Session) -> PolicyDetail:
    """Build a PolicyDetail response with assigned printers."""
    printer_rows = (
        db.query(Printer, Plant.name.label("plant_name"), Area.name.label("area_name"))
        .join(PolicyPrinter, PolicyPrinter.printer_id == Printer.id)
        .outerjoin(Plant, Printer.plant_id == Plant.id)
        .outerjoin(Area, Printer.area_id == Area.id)
        .filter(PolicyPrinter.policy_id == policy.id)
        .all()
    )
    printers = [
        PolicyPrinterItem(
            id=p.id,
            code=p.code,
            serial_number=p.serial_number,
            plant_name=plant_name,
            area_name=area_name,
        )
        for p, plant_name, area_name in printer_rows
    ]
    client = db.get(Client, policy.client_id)
    return PolicyDetail(
        id=policy.id,
        code=policy.code,
        folio=policy.folio,
        client_id=policy.client_id,
        client_name=client.name if client else policy.client_id,
        coverage_type=policy.coverage_type,
        start_date=policy.start_date,
        end_date=policy.end_date,
        status=_policy_status(policy.end_date),
        printer_count=len(printers),
        sla_notes=policy.sla_notes,
        frequency_maintenance=policy.frequency_maintenance,
        printers=printers,
    )


# ---------------------------------------------------------------------------
# POST /api/admin/policies  — Crear póliza
# ---------------------------------------------------------------------------

@router.post("/policies", response_model=dict, status_code=201)
def create_policy(body: PolicyCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new maintenance policy."""
    if not db.get(Client, body.client_id):
        raise HTTPException(status_code=404, detail="Client not found")

    if body.start_date >= body.end_date:
        raise HTTPException(status_code=422, detail="start_date must be before end_date")

    if db.query(Policy).filter(Policy.folio == body.folio).first():
        raise HTTPException(status_code=409, detail=f"Folio '{body.folio}' already exists")

    policy = Policy(
        id=str(uuid.uuid4()),
        code=_next_policy_code(db),
        client_id=body.client_id,
        folio=body.folio,
        start_date=body.start_date,
        end_date=body.end_date,
        coverage_type=body.coverage_type,
        sla_notes=body.sla_notes,
        frequency_maintenance=body.frequency_maintenance,
        status=_policy_status(body.end_date),
    )
    db.add(policy)
    db.commit()
    db.refresh(policy)
    return _policy_detail(policy, db).model_dump()


# ---------------------------------------------------------------------------
# GET /api/admin/policies/next-folio  — Siguiente folio sugerido
# ---------------------------------------------------------------------------

@router.get("/policies/next-folio", response_model=dict)
def get_next_policy_folio(db: Session = Depends(get_db)) -> dict:
    """Return the next suggested folio in POL-YYYY-NNN format for the current year."""
    import re
    year = datetime.now().year
    prefix = f"POL-{year}-"
    rows = db.query(Policy.folio).filter(Policy.folio.like(f"{prefix}%")).all()
    nums = []
    for (folio,) in rows:
        match = re.match(rf"POL-{year}-(\d+)$", folio or "")
        if match:
            try:
                nums.append(int(match.group(1)))
            except ValueError:
                pass
    next_num = max(nums, default=0) + 1
    return {"folio": f"{prefix}{next_num:03d}"}


# ---------------------------------------------------------------------------
# GET /api/admin/policies/{policy_id}  — Detalle
# ---------------------------------------------------------------------------

@router.get("/policies/{policy_id}", response_model=dict)
def get_policy(policy_id: str, db: Session = Depends(get_db)) -> dict:
    """Get policy detail including assigned printers."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    return _policy_detail(policy, db).model_dump()


# ---------------------------------------------------------------------------
# PUT /api/admin/policies/{policy_id}  — Editar
# ---------------------------------------------------------------------------

@router.put("/policies/{policy_id}", response_model=dict)
def update_policy(
    policy_id: str, body: PolicyUpdate, db: Session = Depends(get_db)
) -> dict:
    """Update mutable fields of a policy and recalculate status."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    if body.folio is not None:
        conflict = (
            db.query(Policy)
            .filter(Policy.folio == body.folio, Policy.id != policy_id)
            .first()
        )
        if conflict:
            raise HTTPException(status_code=409, detail=f"Folio '{body.folio}' already exists")
        policy.folio = body.folio

    if body.start_date is not None:
        policy.start_date = body.start_date
    if body.end_date is not None:
        policy.end_date = body.end_date
    if body.coverage_type is not None:
        policy.coverage_type = body.coverage_type
    if body.sla_notes is not None:
        policy.sla_notes = body.sla_notes
    if body.frequency_maintenance is not None:
        policy.frequency_maintenance = body.frequency_maintenance

    if policy.start_date >= policy.end_date:
        raise HTTPException(status_code=422, detail="start_date must be before end_date")

    policy.status = _policy_status(policy.end_date)
    db.commit()
    db.refresh(policy)
    return _policy_detail(policy, db).model_dump()


# ---------------------------------------------------------------------------
# DELETE /api/admin/policies/{policy_id}  — Soft delete
# ---------------------------------------------------------------------------

@router.delete("/policies/{policy_id}", status_code=200)
def delete_policy(policy_id: str, db: Session = Depends(get_db)) -> dict:
    """Soft-delete a policy by setting its status to 'Deleted'."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    if policy.status == "Deleted":
        raise HTTPException(status_code=409, detail="Policy is already deleted")

    policy.status = "Deleted"
    db.commit()
    return {"success": True, "id": policy_id}


# ---------------------------------------------------------------------------
# POST /api/admin/policies/{policy_id}/printers  — Asignar impresoras
# ---------------------------------------------------------------------------

@router.post("/policies/{policy_id}/printers", status_code=200)
def assign_printers(
    policy_id: str, body: AssignPrintersRequest, db: Session = Depends(get_db)
) -> dict:
    """Assign one or more printers to a policy. Validates same client."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    added, skipped, invalid = [], [], []

    for printer_id in body.printer_ids:
        printer = db.get(Printer, printer_id)
        if not printer:
            invalid.append(printer_id)
            continue

        if printer.client_id != policy.client_id:
            raise HTTPException(
                status_code=422,
                detail=f"Printer {printer.serial_number} belongs to a different client",
            )

        already = (
            db.query(PolicyPrinter)
            .filter(
                PolicyPrinter.policy_id == policy_id,
                PolicyPrinter.printer_id == printer_id,
            )
            .first()
        )
        if already:
            skipped.append(printer_id)
            continue

        db.add(PolicyPrinter(id=str(uuid.uuid4()), policy_id=policy_id, printer_id=printer_id))
        added.append(printer_id)

    db.commit()
    return {"success": True, "added": added, "skipped": skipped, "invalid": invalid}


# ---------------------------------------------------------------------------
# DELETE /api/admin/policies/{policy_id}/printers/{printer_id}  — Quitar impresora
# ---------------------------------------------------------------------------

@router.delete("/policies/{policy_id}/printers/{printer_id}", status_code=200)
def remove_printer_from_policy(
    policy_id: str, printer_id: str, db: Session = Depends(get_db)
) -> dict:
    """Remove a printer from a policy."""
    link = (
        db.query(PolicyPrinter)
        .filter(
            PolicyPrinter.policy_id == policy_id,
            PolicyPrinter.printer_id == printer_id,
        )
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="Printer not assigned to this policy")

    db.delete(link)
    db.commit()
    return {"success": True, "policy_id": policy_id, "printer_id": printer_id}


# ===========================================================================
# FASE 7 — Asignaciones técnico-impresora
# ===========================================================================

@router.get("/policies/{policy_id}/assignments", response_model=dict)
def list_assignments(policy_id: str, db: Session = Depends(get_db)) -> dict:
    """List technician assignments for printers in a policy."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    rows = (
        db.query(PolicyPrinterAssignment)
        .filter(PolicyPrinterAssignment.policy_id == policy_id)
        .all()
    )
    items = []
    for a in rows:
        printer = db.get(Printer, a.printer_id)
        tech = db.get(User, a.technician_id)
        items.append(PolicyPrinterAssignmentItem(
            id=a.id,
            policy_id=a.policy_id,
            printer_id=a.printer_id,
            printer_serial=printer.serial_number if printer else None,
            technician_id=a.technician_id,
            technician_name=tech.name if tech else None,
            technician_code=tech.code if tech else None,
            assigned_at=a.assigned_at,
        ).model_dump())
    return {"total": len(items), "items": items}


@router.post("/policies/{policy_id}/assignments", response_model=dict)
def assign_technician(
    policy_id: str, body: AssignTechnicianRequest, db: Session = Depends(get_db)
) -> dict:
    """Upsert: assign a technician to a printer within a policy."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    # Validate printer belongs to policy
    pp = (
        db.query(PolicyPrinter)
        .filter(PolicyPrinter.policy_id == policy_id, PolicyPrinter.printer_id == body.printer_id)
        .first()
    )
    if not pp:
        raise HTTPException(status_code=422, detail="Printer not assigned to this policy")

    if not db.get(User, body.technician_id):
        raise HTTPException(status_code=404, detail="Technician not found")

    existing = (
        db.query(PolicyPrinterAssignment)
        .filter(
            PolicyPrinterAssignment.policy_id == policy_id,
            PolicyPrinterAssignment.printer_id == body.printer_id,
        )
        .first()
    )
    if existing:
        existing.technician_id = body.technician_id
        existing.assigned_at = _now_utc()
    else:
        existing = PolicyPrinterAssignment(
            id=str(uuid.uuid4()),
            policy_id=policy_id,
            printer_id=body.printer_id,
            technician_id=body.technician_id,
            assigned_at=_now_utc(),
        )
        db.add(existing)

    db.commit()
    db.refresh(existing)

    printer = db.get(Printer, body.printer_id)
    tech = db.get(User, body.technician_id)
    return PolicyPrinterAssignmentItem(
        id=existing.id,
        policy_id=existing.policy_id,
        printer_id=existing.printer_id,
        printer_serial=printer.serial_number if printer else None,
        technician_id=existing.technician_id,
        technician_name=tech.name if tech else None,
        technician_code=tech.code if tech else None,
        assigned_at=existing.assigned_at,
    ).model_dump()


@router.delete("/policies/{policy_id}/assignments/{printer_id}", status_code=200)
def delete_assignment(
    policy_id: str, printer_id: str, db: Session = Depends(get_db)
) -> dict:
    """Remove a technician assignment for a printer in a policy."""
    assignment = (
        db.query(PolicyPrinterAssignment)
        .filter(
            PolicyPrinterAssignment.policy_id == policy_id,
            PolicyPrinterAssignment.printer_id == printer_id,
        )
        .first()
    )
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(assignment)
    db.commit()
    return {"success": True, "policy_id": policy_id, "printer_id": printer_id}


# ===========================================================================
# FASE 7 — Entregas de póliza (admin)
# ===========================================================================

def _create_policy_delivery(policy_id: str, body: PolicyDeliveryCreate, db: Session) -> PolicyDeliveryItem:
    """Shared logic: create delivery + link reports + mark as signed."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    delivery = PolicyDelivery(
        id=str(uuid.uuid4()),
        policy_id=policy_id,
        delivery_date=body.delivery_date,
        signature_name=body.signature_name,
        signature_role=body.signature_role,
        tech_id=body.tech_id,
        signature_image_path=body.signature_image_path,
    )
    db.add(delivery)
    db.flush()

    for report_id in body.report_ids:
        db.add(PolicyDeliveryReport(
            id=str(uuid.uuid4()),
            delivery_id=delivery.id,
            report_id=report_id,
        ))
        from app.models.report import Report as ReportModel
        report = db.get(ReportModel, report_id)
        if report:
            report.status = "signed"

    db.commit()
    db.refresh(delivery)

    return PolicyDeliveryItem(
        id=delivery.id,
        policy_id=delivery.policy_id,
        delivery_date=delivery.delivery_date,
        signature_name=delivery.signature_name,
        signature_role=delivery.signature_role,
        tech_id=delivery.tech_id,
        signature_image_path=delivery.signature_image_path,
        report_count=len(body.report_ids),
    )


@router.get("/policies/{policy_id}/deliveries", response_model=dict)
def list_deliveries(policy_id: str, db: Session = Depends(get_db)) -> dict:
    """List delivery history for a policy."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    rows = (
        db.query(PolicyDelivery)
        .filter(PolicyDelivery.policy_id == policy_id)
        .order_by(PolicyDelivery.delivery_date.desc())
        .all()
    )
    items = []
    for d in rows:
        report_count = (
            db.query(func.count(PolicyDeliveryReport.id))
            .filter(PolicyDeliveryReport.delivery_id == d.id)
            .scalar() or 0
        )
        items.append(PolicyDeliveryItem(
            id=d.id,
            policy_id=d.policy_id,
            delivery_date=d.delivery_date,
            signature_name=d.signature_name,
            signature_role=d.signature_role,
            tech_id=d.tech_id,
            signature_image_path=d.signature_image_path,
            report_count=report_count,
        ).model_dump())
    return {"total": len(items), "items": items}


@router.get("/policy-deliveries/{delivery_id}/detail", response_model=dict)
def get_delivery_detail(delivery_id: str, db: Session = Depends(get_db)) -> dict:
    """Get full delivery detail including each report's printer info."""
    delivery = db.get(PolicyDelivery, delivery_id)
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")

    policy = db.get(Policy, delivery.policy_id)
    tech = db.get(User, delivery.tech_id)

    delivery_reports = (
        db.query(PolicyDeliveryReport)
        .filter(PolicyDeliveryReport.delivery_id == delivery_id)
        .all()
    )

    items = []
    for dr in delivery_reports:
        report = db.get(Report, dr.report_id)
        if not report:
            continue
        printer = db.get(Printer, report.printer_id) if report.printer_id else None
        model = db.get(CatalogModel, printer.model_id) if printer else None
        items.append({
            "report_id": dr.report_id,
            "serial_number": printer.serial_number if printer else None,
            "model_name": f"{model.brand} {model.model_name}" if model else None,
            "service_type": report.service_type,
            "service_date": report.service_date.isoformat() if report.service_date else None,
            "status": report.status,
        })

    return {
        "id": delivery.id,
        "policy_id": delivery.policy_id,
        "policy_folio": policy.folio if policy else None,
        "delivery_date": delivery.delivery_date.isoformat(),
        "signature_name": delivery.signature_name,
        "signature_role": delivery.signature_role,
        "tech_id": delivery.tech_id,
        "tech_name": tech.name if tech else None,
        "report_count": len(items),
        "reports": items,
    }


@router.post("/policies/{policy_id}/deliveries", response_model=dict, status_code=201)
def create_delivery_admin(
    policy_id: str, body: PolicyDeliveryCreate, db: Session = Depends(get_db)
) -> dict:
    """Create a delivery for a policy (admin side)."""
    return _create_policy_delivery(policy_id, body, db).model_dump()


# ===========================================================================
# CRUD — Técnicos
# ===========================================================================

@router.post("/technicians", response_model=dict, status_code=201)
def create_technician(body: TechnicianCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new technician user."""
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(status_code=409, detail=f"Email '{body.email}' already exists")

    _validate_password(body.password, body.name)

    tech = User(
        id=str(uuid.uuid4()),
        code=_next_code(db, User, "T"),
        name=body.name,
        email=body.email,
        password_hash=_hash_password(body.password),
        role=body.role,
        is_active=True,
    )
    db.add(tech)
    db.commit()
    db.refresh(tech)
    return TechnicianListItem(
        id=tech.id,
        code=tech.code,
        name=tech.name,
        email=tech.email,
        role=tech.role,
        reports_count=0,
        last_sync_at=None,
    ).model_dump()


@router.put("/technicians/{tech_id}", response_model=dict)
def update_technician(
    tech_id: str, body: TechnicianUpdate, db: Session = Depends(get_db)
) -> dict:
    """Update a technician's mutable fields."""
    tech = db.get(User, tech_id)
    if not tech:
        raise HTTPException(status_code=404, detail="Technician not found")

    if body.email is not None:
        conflict = db.query(User).filter(User.email == body.email, User.id != tech_id).first()
        if conflict:
            raise HTTPException(status_code=409, detail=f"Email '{body.email}' already exists")
        tech.email = body.email
    if body.name is not None:
        tech.name = body.name
    if body.role is not None:
        tech.role = body.role
    if body.password is not None:
        _validate_password(body.password, tech.name)
        tech.password_hash = _hash_password(body.password)

    db.commit()
    db.refresh(tech)

    reports_count = (
        db.query(func.count(Report.id)).filter(Report.tech_id == tech_id).scalar() or 0
    )
    return TechnicianListItem(
        id=tech.id,
        code=tech.code,
        name=tech.name,
        email=tech.email,
        role=tech.role,
        reports_count=reports_count,
        last_sync_at=tech.last_sync_at,
    ).model_dump()


@router.delete("/technicians/{tech_id}", status_code=200)
def delete_technician(tech_id: str, db: Session = Depends(get_db)) -> dict:
    """Soft-delete a technician (set is_active=False)."""
    tech = db.get(User, tech_id)
    if not tech:
        raise HTTPException(status_code=404, detail="Technician not found")
    if not tech.is_active:
        raise HTTPException(status_code=409, detail="Technician is already inactive")

    tech.is_active = False
    db.commit()
    return {"success": True, "id": tech_id}


# ===========================================================================
# CRUD — Clientes
# ===========================================================================

@router.post("/clients", response_model=dict, status_code=201)
def create_client(body: ClientCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new client."""
    client = Client(
        id=str(uuid.uuid4()),
        name=body.name,
        rfc=body.rfc,
        address=body.address,
        is_active=True,
    )
    db.add(client)
    db.commit()
    db.refresh(client)
    return ClientListItem(
        id=client.id,
        name=client.name,
        rfc=client.rfc,
        address=client.address,
        is_active=client.is_active,
        plant_count=0,
        printer_count=0,
        active_policy_count=0,
    ).model_dump()


@router.put("/clients/{client_id}", response_model=dict)
def update_client(
    client_id: str, body: ClientUpdate, db: Session = Depends(get_db)
) -> dict:
    """Update a client's mutable fields."""
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    if body.name is not None:
        client.name = body.name
    if body.rfc is not None:
        client.rfc = body.rfc
    if body.address is not None:
        client.address = body.address

    db.commit()
    db.refresh(client)

    plant_count = db.query(func.count(Plant.id)).filter(Plant.client_id == client_id).scalar() or 0
    printer_count = db.query(func.count(Printer.id)).filter(Printer.client_id == client_id).scalar() or 0
    active_policy_count = (
        db.query(func.count(Policy.id))
        .filter(Policy.client_id == client_id, Policy.status == "Active")
        .scalar() or 0
    )
    return ClientListItem(
        id=client.id,
        name=client.name,
        rfc=client.rfc,
        address=client.address,
        is_active=client.is_active,
        plant_count=plant_count,
        printer_count=printer_count,
        active_policy_count=active_policy_count,
    ).model_dump()


@router.delete("/clients/{client_id}", status_code=200)
def delete_client(client_id: str, db: Session = Depends(get_db)) -> dict:
    """Soft-delete a client (set is_active=False)."""
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if not client.is_active:
        raise HTTPException(status_code=409, detail="Client is already inactive")

    client.is_active = False
    db.commit()
    return {"success": True, "id": client_id}


# ===========================================================================
# Plants & Areas (auxiliares para crear impresoras)
# ===========================================================================

@router.get("/plants", response_model=dict)
def list_plants(
    client_id: str | None = Query(None),
    db: Session = Depends(get_db),
) -> dict:
    """List plants, optionally filtered by client."""
    q = db.query(Plant)
    if client_id:
        q = q.filter(Plant.client_id == client_id)
    rows = q.order_by(Plant.name).all()
    items = [PlantListItem(id=p.id, name=p.name, client_id=p.client_id, contact_name=p.contact_name, phone=p.phone).model_dump() for p in rows]
    return {"total": len(items), "items": items}


@router.post("/plants", response_model=dict, status_code=201)
def create_plant(body: PlantCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new plant for a client."""
    if not db.get(Client, body.client_id):
        raise HTTPException(status_code=404, detail="Client not found")

    plant = Plant(
        id=str(uuid.uuid4()),
        client_id=body.client_id,
        name=body.name,
        contact_name=body.contact_name,
        phone=body.contact_phone,
    )
    db.add(plant)
    db.commit()
    db.refresh(plant)
    return PlantListItem(
        id=plant.id, name=plant.name, client_id=plant.client_id,
        contact_name=plant.contact_name, phone=plant.phone,
    ).model_dump()


@router.put("/plants/{plant_id}", response_model=dict)
def update_plant(plant_id: str, body: PlantUpdate, db: Session = Depends(get_db)) -> dict:
    plant = db.get(Plant, plant_id)
    if not plant:
        raise HTTPException(status_code=404, detail="Plant not found")
    if body.contact_name is not None:
        plant.contact_name = body.contact_name
    if body.contact_phone is not None:
        plant.phone = body.contact_phone
    db.commit()
    db.refresh(plant)
    return PlantListItem(
        id=plant.id, name=plant.name, client_id=plant.client_id,
        contact_name=plant.contact_name, phone=plant.phone,
    ).model_dump()


@router.get("/areas", response_model=dict)
def list_areas(
    plant_id: str | None = Query(None),
    db: Session = Depends(get_db),
) -> dict:
    """List areas, optionally filtered by plant."""
    q = db.query(Area)
    if plant_id:
        q = q.filter(Area.plant_id == plant_id)
    rows = q.order_by(Area.name).all()
    items = [AreaListItem(id=a.id, name=a.name, plant_id=a.plant_id).model_dump() for a in rows]
    return {"total": len(items), "items": items}


@router.post("/areas", response_model=dict, status_code=201)
def create_area(body: AreaCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new area inside a plant."""
    if not db.get(Plant, body.plant_id):
        raise HTTPException(status_code=404, detail="Plant not found")

    area = Area(
        id=str(uuid.uuid4()),
        plant_id=body.plant_id,
        name=body.name,
    )
    db.add(area)
    db.commit()
    db.refresh(area)
    return AreaListItem(id=area.id, name=area.name, plant_id=area.plant_id).model_dump()


@router.get("/catalog/models", response_model=dict)
def list_catalog_models(db: Session = Depends(get_db)) -> dict:
    """List active printer models from the catalog."""
    rows = db.query(CatalogModel).filter(CatalogModel.is_active.is_(True)).order_by(
        CatalogModel.brand, CatalogModel.model_name
    ).all()
    items = [
        CatalogModelItem(id=m.id, brand=m.brand, model_name=m.model_name, dpi=m.dpi).model_dump()
        for m in rows
    ]
    return {"total": len(items), "items": items}


@router.post("/catalog/models", response_model=dict, status_code=201)
def create_catalog_model(body: CatalogModelCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new printer model in the catalog."""
    model = CatalogModel(
        id=str(uuid.uuid4()),
        brand=body.brand,
        model_name=body.model_name,
        dpi=body.dpi,
        is_active=True,
    )
    db.add(model)
    db.commit()
    db.refresh(model)
    return CatalogModelItem(id=model.id, brand=model.brand, model_name=model.model_name, dpi=model.dpi).model_dump()


# ===========================================================================
# CRUD — Impresoras
# ===========================================================================

@router.post("/printers", response_model=dict, status_code=201)
def create_printer(body: PrinterCreate, db: Session = Depends(get_db)) -> dict:
    """Create a new printer."""
    if not db.get(Client, body.client_id):
        raise HTTPException(status_code=404, detail="Client not found")
    if not db.get(Plant, body.plant_id):
        raise HTTPException(status_code=404, detail="Plant not found")
    if not db.get(Area, body.area_id):
        raise HTTPException(status_code=404, detail="Area not found")
    if not db.get(CatalogModel, body.model_id):
        raise HTTPException(status_code=404, detail="Model not found")

    if db.query(Printer).filter(Printer.serial_number == body.serial_number).first():
        raise HTTPException(status_code=409, detail=f"Serial '{body.serial_number}' already exists")

    printer = Printer(
        id=str(uuid.uuid4()),
        code=_next_code(db, Printer, "I"),
        qr_uuid=body.qr_uuid or str(uuid.uuid4()),
        serial_number=body.serial_number,
        client_id=body.client_id,
        plant_id=body.plant_id,
        area_id=body.area_id,
        model_id=body.model_id,
        is_active=True,
    )
    db.add(printer)
    db.commit()
    db.refresh(printer)

    plant = db.get(Plant, body.plant_id)
    area = db.get(Area, body.area_id)
    client = db.get(Client, body.client_id)
    catalog_model = db.get(CatalogModel, body.model_id)
    return PrinterListItem(
        id=printer.id,
        code=printer.code,
        serial_number=printer.serial_number,
        client_name=client.name if client else None,
        plant_name=plant.name if plant else None,
        area_name=area.name if area else None,
        model_brand=catalog_model.brand if catalog_model else None,
        model_name=catalog_model.model_name if catalog_model else None,
        model_dpi=catalog_model.dpi if catalog_model else None,
        last_service_date=None,
        printer_status="Sin Historial",
    ).model_dump()


@router.put("/printers/{printer_id}", response_model=dict)
def update_printer(
    printer_id: str, body: PrinterUpdate, db: Session = Depends(get_db)
) -> dict:
    """Update a printer's mutable fields."""
    printer = db.get(Printer, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    if body.serial_number is not None:
        conflict = (
            db.query(Printer)
            .filter(Printer.serial_number == body.serial_number, Printer.id != printer_id)
            .first()
        )
        if conflict:
            raise HTTPException(status_code=409, detail=f"Serial '{body.serial_number}' already exists")
        printer.serial_number = body.serial_number
    if body.client_id is not None:
        if not db.get(Client, body.client_id):
            raise HTTPException(status_code=404, detail="Client not found")
        printer.client_id = body.client_id
    if body.plant_id is not None:
        if not db.get(Plant, body.plant_id):
            raise HTTPException(status_code=404, detail="Plant not found")
        printer.plant_id = body.plant_id
    if body.area_id is not None:
        if not db.get(Area, body.area_id):
            raise HTTPException(status_code=404, detail="Area not found")
        printer.area_id = body.area_id
    if body.model_id is not None:
        if not db.get(CatalogModel, body.model_id):
            raise HTTPException(status_code=404, detail="Model not found")
        printer.model_id = body.model_id

    db.commit()
    db.refresh(printer)

    plant = db.get(Plant, printer.plant_id)
    area = db.get(Area, printer.area_id)
    client = db.get(Client, printer.client_id)
    catalog_model = db.get(CatalogModel, printer.model_id)
    return PrinterListItem(
        id=printer.id,
        code=printer.code,
        serial_number=printer.serial_number,
        client_name=client.name if client else None,
        plant_name=plant.name if plant else None,
        area_name=area.name if area else None,
        model_brand=catalog_model.brand if catalog_model else None,
        model_name=catalog_model.model_name if catalog_model else None,
        last_service_date=None,
        printer_status="Sin Historial",
    ).model_dump()


@router.delete("/printers/{printer_id}", status_code=200)
def delete_printer(printer_id: str, db: Session = Depends(get_db)) -> dict:
    """Soft-delete a printer (set is_active=False)."""
    printer = db.get(Printer, printer_id)
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")
    if not printer.is_active:
        raise HTTPException(status_code=409, detail="Printer is already inactive")

    printer.is_active = False
    db.commit()
    return {"success": True, "id": printer_id}


# ---------------------------------------------------------------------------
# Policy Visits
# ---------------------------------------------------------------------------

_FREQUENCY_VISIT_COUNT: dict[str, int] = {
    "Mensual": 12,
    "Bimestral": 6,
    "Trimestral": 4,
    "Semestral": 2,
    "Anual": 1,
}


def _build_visit_item(visit: "PolicyVisit", db: Session) -> PolicyVisitItem:
    """Build a PolicyVisitItem with computed counts."""
    total_printers = (
        db.query(func.count(PolicyPrinter.id))
        .filter(PolicyPrinter.policy_id == visit.policy_id)
        .scalar()
        or 0
    )
    # Count reports in pending_delivery status for this policy's printers
    printer_ids = [
        pp.printer_id
        for pp in db.query(PolicyPrinter)
        .filter(PolicyPrinter.policy_id == visit.policy_id)
        .all()
    ]
    from app.models.report import Report as ReportModel
    attended_count = 0
    if printer_ids:
        attended_count = (
            db.query(func.count(ReportModel.id))
            .filter(
                ReportModel.printer_id.in_(printer_ids),
                ReportModel.status == "pending_delivery",
            )
            .scalar()
            or 0
        )
    return PolicyVisitItem(
        id=visit.id,
        policy_id=visit.policy_id,
        visit_number=visit.visit_number,
        scheduled_date=visit.scheduled_date.isoformat() if visit.scheduled_date else None,
        status=visit.status,
        started_at=visit.started_at,
        completed_at=visit.completed_at,
        created_at=visit.created_at,
        attended_count=attended_count,
        total_printers=total_printers,
    )


@router.get("/policies/{policy_id}/visits", response_model=list)
def list_policy_visits(policy_id: str, db: Session = Depends(get_db)) -> list:
    """List all visits for a given policy, ordered by visit_number."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    visits = (
        db.query(PolicyVisit)
        .filter(PolicyVisit.policy_id == policy_id)
        .order_by(PolicyVisit.visit_number)
        .all()
    )
    return [_build_visit_item(v, db).model_dump() for v in visits]


@router.post("/policies/{policy_id}/visits/generate", response_model=list, status_code=201)
def generate_policy_visits(
    policy_id: str,
    _body: GenerateVisitsRequest,
    db: Session = Depends(get_db),
) -> list:
    """Auto-generate N visits based on policy.frequency_maintenance.

    Only allowed if the policy has no visits yet.
    Schedules dates evenly between start_date and end_date.
    """
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    existing = (
        db.query(PolicyVisit).filter(PolicyVisit.policy_id == policy_id).count()
    )
    if existing > 0:
        raise HTTPException(
            status_code=409,
            detail="Policy already has visits. Delete them before regenerating.",
        )

    # frequency_maintenance may be stored as "Trimestral (4 visitas)" — extract the keyword before "("
    raw_freq = (policy.frequency_maintenance or "Anual").split("(")[0].strip().title()
    n_visits = _FREQUENCY_VISIT_COUNT.get(raw_freq, 1)
    print(f"[Visits] policy.frequency_maintenance raw: '{policy.frequency_maintenance}'")
    print(f"[Visits] freq normalizado: '{raw_freq}'")
    print(f"[Visits] n_visits calculado: {n_visits}")

    from datetime import date, timedelta
    start: date = policy.start_date.date() if hasattr(policy.start_date, "date") else policy.start_date
    end: date = policy.end_date.date() if hasattr(policy.end_date, "date") else policy.end_date
    total_days = (end - start).days

    created_visits: list[PolicyVisit] = []
    for i in range(n_visits):
        # Distribute evenly between start_date and end_date (endpoint-inclusive).
        # For a single visit, place it on start_date.
        if n_visits == 1:
            offset_days = 0
        else:
            offset_days = round(total_days * i / (n_visits - 1))
        scheduled = start + timedelta(days=offset_days)

        visit = PolicyVisit(
            id=str(uuid.uuid4()),
            policy_id=policy_id,
            visit_number=i + 1,
            scheduled_date=scheduled,
            status="scheduled",
        )
        db.add(visit)
        created_visits.append(visit)

    db.commit()
    for v in created_visits:
        db.refresh(v)

    return [_build_visit_item(v, db).model_dump() for v in created_visits]


@router.patch("/policies/{policy_id}/visits/{visit_id}", response_model=dict)
def update_policy_visit(
    policy_id: str,
    visit_id: str,
    body: PolicyVisitUpdate,
    db: Session = Depends(get_db),
) -> dict:
    """Update visit status.

    Validates: if status=in_progress, no other visit of this policy may already
    be in_progress.
    """
    visit = db.get(PolicyVisit, visit_id)
    if not visit or visit.policy_id != policy_id:
        raise HTTPException(status_code=404, detail="Visit not found")

    if body.status == "in_progress" and visit.status != "in_progress":
        # Check for existing in_progress visit on the same policy
        conflict = (
            db.query(PolicyVisit)
            .filter(
                PolicyVisit.policy_id == policy_id,
                PolicyVisit.status == "in_progress",
                PolicyVisit.id != visit_id,
            )
            .first()
        )
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Visit {conflict.visit_number} is already in_progress for this policy.",
            )
        visit.started_at = _now_utc()

    if body.status == "completed" and visit.status != "completed":
        visit.completed_at = _now_utc()

    visit.status = body.status
    if body.scheduled_date is not None:
        from datetime import date as date_type
        visit.scheduled_date = date_type.fromisoformat(body.scheduled_date)

    db.commit()
    db.refresh(visit)
    return _build_visit_item(visit, db).model_dump()


@router.delete("/policies/{policy_id}/visits", status_code=200)
def delete_all_policy_visits(
    policy_id: str,
    db: Session = Depends(get_db),
) -> dict:
    """Delete ALL visits for a policy at once. Returns count of deleted visits."""
    policy = db.get(Policy, policy_id)
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")

    deleted = (
        db.query(PolicyVisit)
        .filter(PolicyVisit.policy_id == policy_id)
        .delete(synchronize_session=False)
    )
    db.commit()
    print(f"[DELETE all visits] policy_id={policy_id} deleted={deleted}")
    return {"deleted": deleted}


@router.delete("/policies/{policy_id}/visits/{visit_id}", status_code=204)
def delete_policy_visit(
    policy_id: str,
    visit_id: str,
    db: Session = Depends(get_db),
) -> None:
    """Delete a visit regardless of status. Admin has full control."""
    print(f"[DELETE visit] policy_id={policy_id} visit_id={visit_id}")
    visit = db.query(PolicyVisit).filter(
        PolicyVisit.id == visit_id,
        PolicyVisit.policy_id == policy_id,
    ).first()
    print(f"[DELETE visit] encontrada: {visit is not None}")
    if not visit:
        raise HTTPException(status_code=404, detail="Visit not found")
    db.delete(visit)
    db.commit()


# ===========================================================================
# Dashboard — endpoints de resumen rápido
# ===========================================================================

@router.get("/dashboard/reports-by-day", response_model=list)
def dashboard_reports_by_day(db: Session = Depends(get_db)) -> list:
    """Return last 7 days of reports grouped by service type. Always returns all 7 days."""
    today = date.today()
    days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    start_dt = datetime.combine(days[0], time.min)
    end_dt = datetime.combine(today, time.max)

    rows = (
        db.query(
            func.date(Report.service_date).label("fecha"),
            Report.service_type,
            func.count(Report.id).label("total"),
        )
        .filter(Report.service_date >= start_dt, Report.service_date <= end_dt)
        .group_by(func.date(Report.service_date), Report.service_type)
        .all()
    )

    # Build a lookup: date_str → { service_type: count }
    data_map: dict[str, dict[str, int]] = {}
    for fecha, service_type, total in rows:
        fecha_str = str(fecha)
        if fecha_str not in data_map:
            data_map[fecha_str] = {}
        data_map[fecha_str][service_type or ""] = total

    result = []
    for d in days:
        fecha_str = str(d)
        counts = data_map.get(fecha_str, {})
        result.append({
            "fecha": fecha_str,
            "total": sum(counts.values()),
            "preventivos": counts.get("Preventivo", 0),
            "correctivos": counts.get("Correctivo", 0),
            "diagnosticos": counts.get("Diagnóstico", 0),
        })
    return result


@router.get("/dashboard/printers-attention", response_model=list)
def dashboard_printers_attention(db: Session = Depends(get_db)) -> list:
    """Return up to 5 active printers whose latest report has at least one damage key true."""
    latest_report_sq = (
        db.query(
            Report.printer_id,
            func.max(Report.service_date).label("max_date"),
        )
        .group_by(Report.printer_id)
        .subquery()
    )
    latest_full_sq = (
        db.query(Report.printer_id, Report.technical_checkboxes)
        .join(
            latest_report_sq,
            (Report.printer_id == latest_report_sq.c.printer_id)
            & (Report.service_date == latest_report_sq.c.max_date),
        )
        .subquery()
    )

    rows = (
        db.query(
            Printer,
            Client.name.label("client_name"),
            CatalogModel.brand.label("model_brand"),
            CatalogModel.model_name.label("model_name"),
            latest_full_sq.c.technical_checkboxes.label("last_checkboxes"),
        )
        .outerjoin(Client, Printer.client_id == Client.id)
        .outerjoin(CatalogModel, Printer.model_id == CatalogModel.id)
        .join(latest_full_sq, Printer.id == latest_full_sq.c.printer_id)
        .filter(Printer.is_active.is_(True))
        .all()
    )

    result = []
    for printer, client_name, model_brand, model_name, last_checkboxes in rows:
        if not last_checkboxes:
            continue
        try:
            checkboxes = json.loads(last_checkboxes)
            advertencias = [k for k in _DAMAGE_KEYS if checkboxes.get(k) is True]
        except (json.JSONDecodeError, AttributeError):
            continue
        if not advertencias:
            continue

        model_full = " ".join(filter(None, [model_brand, model_name])) or None
        result.append({
            "id": printer.id,
            "code": printer.code,
            "serial_number": printer.serial_number,
            "model_name": model_full,
            "client_name": client_name,
            "advertencias": advertencias,
        })
        if len(result) >= 5:
            break

    return result


@router.get("/dashboard/policies-expiring", response_model=list)
def dashboard_policies_expiring(db: Session = Depends(get_db)) -> list:
    """Return up to 5 policies expiring within the next 30 days (not Deleted)."""
    now = _now_utc()
    thirty_days_later = now + timedelta(days=30)

    rows = (
        db.query(Policy, Client.name.label("client_name"))
        .join(Client, Policy.client_id == Client.id)
        .filter(
            Policy.end_date >= now,
            Policy.end_date <= thirty_days_later,
            Policy.status != "Deleted",
        )
        .order_by(Policy.end_date)
        .limit(5)
        .all()
    )

    result = []
    for policy, client_name in rows:
        end = policy.end_date.replace(tzinfo=None)
        dias_restantes = max(0, (end.date() - now.date()).days)
        result.append({
            "id": policy.id,
            "folio": policy.folio,
            "client_name": client_name,
            "end_date": end.date().isoformat(),
            "dias_restantes": dias_restantes,
        })
    return result
